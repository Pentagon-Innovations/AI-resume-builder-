import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

# Sample Data
data = {
    "ID": [101, 102, 103, 104, 105],
    "Name": ["Alice", "Bob", "Charlie", "David", "Emma"],
    "Department": ["HR", "IT", "Finance", "Marketing", "IT"],
    "Salary": [50000, 70000, 60000, 55000, 75000],
    "Joining Date": ["2022-06-10", "2021-08-15", "2020-01-25", "2019-11-30", "2023-03-20"]
}

# Create a DataFrame
df = pd.DataFrame(data)

# Save to Excel
file_name = "formatted_data.xlsx"
df.to_excel(file_name, index=False, sheet_name="Employees")

# Load the Workbook
wb = load_workbook(file_name)
ws = wb.active

# Format Headers
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

# Apply formatting to headers
for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(df.columns)):
    for cell in col:
        cell.font = header_font
        cell.fill = header_fill

# Add Borders & Row Formatting
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
alt_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.border = thin_border
    if row[0].row % 2 == 0:  # Apply alternating color
        for cell in row:
            cell.fill = alt_fill

# Auto-adjust column width
for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter  # Get column letter
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = max_length + 2

# Save the Workbook
wb.save(file_name)

print(f"Excel file '{file_name}' created with formatting!")
